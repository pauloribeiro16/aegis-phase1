"""NIST CSF 2.0 Reference Tool (xlsx) parser.

Source: ``csf2.xlsx`` (NIST CSF 2.0 Reference Tool export).

The workbook has 2 sheets:

1. ``Introduction`` — Title, Read Me, Change Log, Generated Date.
2. ``CSF 2.0`` — Core table with 5 columns:

    | Function | Category | Subcategory | Implementation Examples | Informative References |

The Function and Category columns use "merged-cell" semantics — the value
appears only on the first row of each block; subsequent rows leave the cell
empty. Subcategory has one row per item. Implementation Examples and
Informative References cells contain **newline-separated** lists of items.

Withdrawn subcategories (e.g. ``PR.DS-03: [Withdrawn: ...]``) are still
present in the source — we mark them with ``withdrawn: true`` and capture
the original text verbatim.

All output is **structured** — no raw_md dumping. The per-subcategory
shard has every field of the source mapped to a typed JSON field, plus:

- ``implementation_examples``: list of {label, text}
- ``informative_references``: list of {raw, family, ref}
- ``withdrawn``: bool
- ``withdrawal_note``: parsed from the "[Withdrawn: ...]" prefix when present
- ``reference_families``: distinct families used (sorted, e.g. ["CSF v1.1",
  "ISO/IEC 27001:2022", "SP 800-53 Rev 5.1.1"])
"""

from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path
from typing import Any

import openpyxl

# Subcategory id pattern: e.g. "GV.OC-01", "PR.AA-05", "RC.IM-02"
_SUBCAT_ID_RE = re.compile(r"^([A-Z]{2})\.([A-Z]{2})-(\d{2})$")
# Withdrawn pattern: "PR.DS-03: [Withdrawn: Incorporated into ID.AM-08, PR.PS-03]"
_WITHDRAWN_RE = re.compile(r"^\[Withdrawn:\s*(.*?)\]\s*$")
# Implementation Example pattern: "Ex1: ...", "Ex2: ..."
_EXAMPLE_RE = re.compile(r"^Ex(\d+):\s*(.*)$")

# Function code → human name
_FUNCTION_NAME: dict[str, str] = {
    "GV": "Govern",
    "ID": "Identify",
    "PR": "Protect",
    "DE": "Detect",
    "RS": "Respond",
    "RC": "Recover",
}


# ─── family clustering ──────────────────────────────────────────────────


def _family_of(ref: str) -> str:
    """Bucket a reference string into its family.

    The xlsx uses these prefixes (see _REF_FAMILIES for the canonical list).
    Anything not matched falls back to the first token before the first ':'.
    """
    # Order matters: more specific prefixes first
    for prefix, family in _REF_FAMILIES:
        if ref.startswith(prefix):
            return family
    return ref.split(":")[0].strip() or "Other"


# Canonical family table. Each entry is (prefix, family_name). The first
# match wins. Order is significant: longest/most-specific first.
_REF_FAMILIES: list[tuple[str, str]] = [
    ("SP 800-221A", "SP 800-221A"),
    ("SP 800-171", "SP 800-171"),
    ("SP 800-161", "SP 800-161"),
    ("SP 800-61", "SP 800-61"),
    ("SP 800-53", "SP 800-53"),
    ("SP 800-37", "SP 800-37"),
    ("SP 800-30", "SP 800-30"),
    ("SP 800-218", "SP 800-218"),
    ("SP 800-66", "SP 800-66"),
    ("SP-800-37", "SP 800-37"),
    ("SP 800", "SP 800"),  # generic SP 800 fallback
    ("ISO/IEC 27001", "ISO/IEC 27001"),
    ("ISO/IEC 27002", "ISO/IEC 27002"),
    ("ISO/IEC 27701", "ISO/IEC 27701"),
    ("ISO/IEC", "ISO/IEC"),
    ("NICE Framework", "NICE Framework"),
    ("CCMv4.0", "CCM v4.0"),
    ("CCMv3", "CCM v3"),
    ("CCM", "CCM"),
    ("CRI Profile v2.0", "CRI Profile v2.0"),
    ("CRI Profile", "CRI Profile"),
    ("CSF v1.1", "CSF v1.1"),
    ("CSF v1", "CSF v1.x"),
    ("CSF", "CSF"),
    ("CIS Controls v8", "CIS Controls v8"),
    ("CIS Controls v7", "CIS Controls v7"),
    ("CIS Controls", "CIS Controls"),
    ("OWASP Top 10 LLM", "OWASP LLM Top 10"),
    ("OWASP", "OWASP"),
    ("PCI DSS", "PCI DSS"),
    ("AI-SOC", "AI-SOC"),
    ("Guardian-SDK", "Guardian-SDK"),
    ("SDOS", "SDOS"),
    ("SCF", "SCF"),
    ("SSDF", "SSDF"),
    ("CoP", "CoP"),
    ("IRP", "IRP"),
    ("BXAIOS", "BXAIOS"),
]


# ─── cell helpers ───────────────────────────────────────────────────────


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    return bool(isinstance(v, str) and not v.strip())


def _parse_withdrawal(text: str) -> tuple[bool, str | None]:
    """Return (is_withdrawn, withdrawal_note) from a subcategory text cell."""
    m = _WITHDRAWN_RE.match(text.strip())
    if m:
        return True, m.group(1).strip()
    return False, None


def _parse_subcategory_text(
    sc_raw: str,
) -> tuple[str, bool, str | None]:
    """Split the ``Subcategory`` cell into id + text.

    The cell looks like ``GV.OC-01: The organizational mission is understood...``
    or ``PR.DS-03: [Withdrawn: Incorporated into ID.AM-08, PR.PS-03]``.

    Returns ``(id, is_withdrawn, withdrawal_note)`` and the text part (after
    the first ``:``) is returned as the third element via a separate helper
    for the caller to use.
    """
    parts = sc_raw.split(":", 1)
    if len(parts) != 2:
        return sc_raw.strip(), False, None
    sub_id = parts[0].strip()
    rest = parts[1].strip()
    is_wd, wd_note = _parse_withdrawal(rest)
    return sub_id, is_wd, wd_note


def _parse_subcategory_body(sc_raw: str) -> str:
    """Return the text part of a Subcategory cell (after the ID and the ``:``)."""
    parts = sc_raw.split(":", 1)
    if len(parts) != 2:
        return sc_raw.strip()
    rest = parts[1].strip()
    # If the body is a [Withdrawn: ...] note, surface that as-is
    return rest


def _parse_examples(examples_raw: str | None) -> list[dict[str, str]]:
    """Split Implementation Examples cell into list of {label, text}."""
    out: list[dict[str, str]] = []
    if not examples_raw or not examples_raw.strip():
        return out
    for line in examples_raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = _EXAMPLE_RE.match(line)
        if m:
            out.append({"label": f"Ex{m.group(1)}", "text": m.group(2).strip()})
        else:
            # No ExN: prefix — keep as-is with auto-label
            out.append({"label": "Ex?", "text": line})
    return out


def _parse_references(refs_raw: str | None) -> list[dict[str, str]]:
    """Split Informative References cell into list of {raw, family, ref}."""
    out: list[dict[str, str]] = []
    if not refs_raw or not refs_raw.strip():
        return out
    for line in refs_raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        family = _family_of(line)
        # The "ref" part is the substring after the first ':' when present
        ref_id = line.split(":", 1)[1].strip() if ":" in line else line
        out.append({"raw": line, "family": family, "ref": ref_id})
    return out


def _category_id_from_text(cat_text: str) -> str | None:
    """Pull the FUNC.CAT code out of a category text like
    ``Organizational Context (GV.OC): The circumstances ...``"""
    m = re.search(r"\(([A-Z]{2}\.[A-Z]{2})\)", cat_text)
    return m.group(1) if m else None


def _category_name_only(cat_text: str) -> str:
    """Strip the trailing ``: description`` and the ``(FUNC.CAT)`` from a
    category cell. Returns just the name, e.g. ``Organizational Context``."""
    # Remove the (FUNC.CAT) code
    txt = re.sub(r"\s*\([A-Z]{2}\.[A-Z]{2}\)\s*", " ", cat_text)
    # Remove the trailing ": description" (everything after the LAST colon)
    if ":" in txt:
        txt = txt.rsplit(":", 1)[0]
    return txt.strip()


# ─── Introduction sheet ─────────────────────────────────────────────────


def parse_introduction(xlsx_path: Path) -> dict[str, Any]:
    """Parse the ``Introduction`` sheet → title/read_me/change_log/generated_date."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Introduction" not in wb.sheetnames:
        return {}
    ws = wb["Introduction"]
    out: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        key = (row[0] or "").strip() if isinstance(row[0], str) else ""
        val = row[1]
        if not key:
            continue
        if key == "Title":
            out["title"] = str(val).strip() if val else ""
        elif key == "Read Me":
            out["read_me"] = str(val).strip() if val else ""
        elif key == "Change Log":
            out["change_log"] = str(val).strip() if val else ""
        elif key == "Generated Date":
            if isinstance(val, _dt.datetime | _dt.date):
                out["generated_date"] = val.isoformat()
            else:
                out["generated_date"] = str(val).strip() if val else ""
    return out


# ─── CSF 2.0 sheet ─────────────────────────────────────────────────────


def _extract_subcategory_rows(
    ws: Any,
) -> tuple[list[dict[str, Any]], dict[str, str], dict[str, str]]:
    """Walk the CSF 2.0 sheet and extract one record per subcategory.

    Returns ``(rows, function_summary, category_name_by_id)`` where:
      - ``rows`` is a list of dicts (one per subcategory) with keys:
        id, function, function_name, function_summary, category_id,
        category_name, number, title, implementation_examples,
        informative_references, withdrawn, withdrawal_note,
        reference_families, source_locus (xlsx row number, 1-indexed).
      - ``function_summary[function_id]`` = the long Function cell text
        (e.g. ``GOVERN (GV): The organization's cybersecurity risk management...``).
      - ``category_name_by_id[cat_id]`` = the category cell text
        (e.g. ``Organizational Context (GV.OC): The circumstances ...``).

    Merged-cell semantics: the Function and Category cells are populated
    only on the first row of each block; we propagate them downward until
    a new value appears.
    """
    rows: list[dict[str, Any]] = []
    function_summary: dict[str, str] = {}
    category_by_id: dict[str, str] = {}

    current_function_id: str | None = None
    current_function_text: str | None = None
    current_category_id: str | None = None
    current_category_text: str | None = None

    # Header row is row 2 (row 1 is the banner)
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx < 3:
            continue
        col_a, col_b, col_c, col_d, col_e = (list(row) + [None] * 5)[:5]

        # Propagate merged cells: Function (col A) and Category (col B)
        if not _is_blank(col_a):
            current_function_text = str(col_a).strip()
            m = re.match(r"^([A-Z]{2})", current_function_text)
            if m:
                current_function_id = m.group(1)
                if current_function_id not in function_summary:
                    function_summary[current_function_id] = current_function_text

        if not _is_blank(col_b):
            current_category_text = str(col_b).strip()
            current_category_id = _category_id_from_text(current_category_text)
            if current_category_id and current_category_id not in category_by_id:
                category_by_id[current_category_id] = current_category_text

        # Skip if no subcategory cell
        if _is_blank(col_c):
            continue

        sc_raw = str(col_c).strip()
        # Subcategory row
        sub_id, is_wd, wd_note = _parse_subcategory_text(sc_raw)
        m_id = _SUBCAT_ID_RE.match(sub_id)
        if not m_id:
            # Footer / continuation rows (e.g. "RECOVER (RC)" at end) — skip
            continue
        fn_id, cat_id, number = m_id.group(1), m_id.group(2), m_id.group(3)
        sc_body = _parse_subcategory_body(sc_raw)

        examples = _parse_examples(col_d if not _is_blank(col_d) else None)
        references = _parse_references(col_e if not _is_blank(col_e) else None)
        ref_families = sorted({r["family"] for r in references})

        rows.append(
            {
                "id": sub_id,
                "function": fn_id,
                "function_name": _FUNCTION_NAME.get(fn_id, fn_id),
                "function_summary_text": current_function_text or "",
                "category_id": cat_id,
                "category_name_text": current_category_text or "",
                "category_id_resolved": current_category_id or f"{fn_id}.{cat_id}",
                "number": number,
                "title": sc_body,
                "implementation_examples": examples,
                "informative_references": references,
                "reference_families": ref_families,
                "withdrawn": is_wd,
                "withdrawal_note": wd_note,
                "source_locus": {
                    "xlsx_row": row_idx,
                    "function_cell": col_a if not _is_blank(col_a) else None,
                    "category_cell": col_b if not _is_blank(col_b) else None,
                },
            }
        )
    return rows, function_summary, category_by_id


def parse_csf2(xlsx_path: Path) -> dict[str, Any]:
    """Parse csf2.xlsx → top-level dict with introduction + per-subcategory rows.

    Top-level keys:
      - ``introduction`` — title, read_me, change_log, generated_date
      - ``subcategories`` — list of per-subcategory dicts (see _extract_subcategory_rows)
      - ``categories`` — list of {id, function, name, full_text}
      - ``functions`` — list of {id, name, summary_text}
      - ``reference_families`` — list of {family, count, distinct_count,
        example} (sorted by count desc)
      - ``withdrawn_subcategories`` — list of {id, withdrawal_note}
      - ``counts`` — {subcategories, categories, functions, withdrawn,
        reference_cells, reference_distinct_strings, reference_families}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    intro = parse_introduction(xlsx_path)

    if "CSF 2.0" not in wb.sheetnames:
        raise ValueError(f"csf2.xlsx missing 'CSF 2.0' sheet (found {wb.sheetnames})")
    ws = wb["CSF 2.0"]

    subcats, function_summary, category_by_id = _extract_subcategory_rows(ws)

    # Build per-function and per-category aggregations
    functions: list[dict[str, Any]] = []
    seen_fn: dict[str, dict[str, Any]] = {}
    for sc in subcats:
        fid = sc["function"]
        if fid not in seen_fn:
            seen_fn[fid] = {
                "id": fid,
                "name": sc["function_name"],
                "summary_text": sc["function_summary_text"],
                "category_count": 0,
                "subcategory_count": 0,
                "withdrawn_count": 0,
            }
        seen_fn[fid]["subcategory_count"] += 1
        if sc["withdrawn"]:
            seen_fn[fid]["withdrawn_count"] += 1
    # Count categories per function
    fn_cats: dict[str, set[str]] = {}
    for sc in subcats:
        fn_cats.setdefault(sc["function"], set()).add(sc["category_id_resolved"])
    for fid, info in seen_fn.items():
        info["category_count"] = len(fn_cats.get(fid, set()))
    functions = sorted(seen_fn.values(), key=lambda f: f["id"])

    categories: list[dict[str, Any]] = []
    seen_cat: dict[str, dict[str, Any]] = {}
    for sc in subcats:
        cid = sc["category_id_resolved"]
        if cid not in seen_cat:
            seen_cat[cid] = {
                "id": cid,
                "function": sc["function"],
                "function_name": sc["function_name"],
                "name": _category_name_only(sc["category_name_text"]) or sc["category_name_text"],
                "full_text": sc["category_name_text"],
                "subcategory_count": 0,
                "withdrawn_count": 0,
            }
        seen_cat[cid]["subcategory_count"] += 1
        if sc["withdrawn"]:
            seen_cat[cid]["withdrawn_count"] += 1
    categories = sorted(seen_cat.values(), key=lambda c: c["id"])

    # Reference families
    from collections import Counter

    all_refs: list[str] = []
    for sc in subcats:
        for ref in sc["informative_references"]:
            all_refs.append(ref["raw"])
    family_counts: Counter[str] = Counter()
    family_distinct: dict[str, set[str]] = {}
    family_example: dict[str, str] = {}
    for r in all_refs:
        fam = _family_of(r)
        family_counts[fam] += 1
        family_distinct.setdefault(fam, set()).add(r)
        family_example.setdefault(fam, r)
    reference_families: list[dict[str, Any]] = sorted(
        (
            {
                "family": fam,
                "count": family_counts[fam],
                "distinct_count": len(family_distinct[fam]),
                "example": family_example[fam],
            }
            for fam in family_counts
        ),
        key=lambda x: -x["count"],
    )

    # Withdrawn
    withdrawn = [
        {"id": sc["id"], "withdrawal_note": sc["withdrawal_note"]}
        for sc in subcats
        if sc["withdrawn"]
    ]

    return {
        "introduction": intro,
        "subcategories": subcats,
        "categories": categories,
        "functions": functions,
        "reference_families": reference_families,
        "withdrawn_subcategories": withdrawn,
        "counts": {
            "subcategories": len(subcats),
            "categories": len(categories),
            "functions": len(functions),
            "withdrawn": len(withdrawn),
            "reference_cells": len(all_refs),
            "reference_distinct_strings": len(set(all_refs)),
            "reference_families": len(reference_families),
        },
    }


# ─── Per-subcategory shard (for entities/csfs/<ID>.json) ───────────────


def build_shard(
    sc: dict[str, Any],
    intro: dict[str, Any],
    xlsx_path: Path,
) -> dict[str, Any]:
    """Build the per-subcategory entity shard (no raw_md)."""
    return {
        "schema_version": "1.3",
        "kind": "csf",
        "source": "csf2.xlsx",
        "source_path": str(xlsx_path),
        "doc_id": "AEGIS-PREPROC-CSF-2.0-REFTOOL",
        "tool_metadata": {
            "title": intro.get("title", ""),
            "change_log": intro.get("change_log", ""),
            "generated_date": intro.get("generated_date", ""),
        },
        "id": sc["id"],
        "function": sc["function"],
        "function_name": sc["function_name"],
        "function_summary": sc["function_summary_text"],
        "category_id": sc["category_id_resolved"],
        "category_name": _category_name_only(sc["category_name_text"]) or sc["category_name_text"],
        "category_full_text": sc["category_name_text"],
        "number": sc["number"],
        "title": sc["title"],
        "withdrawn": sc["withdrawn"],
        "withdrawal_note": sc["withdrawal_note"],
        "implementation_examples": sc["implementation_examples"],
        "informative_references": sc["informative_references"],
        "reference_families": sc["reference_families"],
        "source_locus": {
            "xlsx_row": sc["source_locus"]["xlsx_row"],
            "sheet": "CSF 2.0",
        },
    }
