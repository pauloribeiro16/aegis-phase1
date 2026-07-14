"""xlsx_generator — render Case_01_Phase1.xlsx via openpyxl.

Produces a 7-sheet workbook (COVER, COMPANY, REGULATIONS, CLAUSES,
COVERAGE, PROPORTIONALITY, GATE) aggregating every Phase 1 artefact.
Each sheet carries a header row, auto-sized columns, frozen header,
and a light zebra style applied conditionally. Cells with no data
fall back to ``-`` so the file is always openable in Excel.
"""

from __future__ import annotations

import logging
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from aegis_phase1.v2.output._common import safe_get

logger = logging.getLogger(__name__)

_FILENAME = "Case_01_Phase1.xlsx"

_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
_ZEBRA_FILL = PatternFill("solid", fgColor="FFF2F2F2")


def generate_xlsx(state: dict[str, Any], output_dir: str) -> dict[str, str]:
    """Generate the consolidated Phase 1 workbook.

    Args:
        state: Pipeline state.
        output_dir: Directory in which the workbook is written.

    Returns:
        Mapping ``AEGIS-P1-XLSX`` -> absolute workbook path.
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _build_cover(wb, state)
    _build_company(wb, state)
    _build_regulations(wb, state)
    _build_clauses(wb, state)
    _build_coverage(wb, state)
    _build_proportionality(wb, state)
    _build_gate(wb, state)

    out_path = Path(output_dir) / _FILENAME
    if out_path.exists():
        versions_dir = Path(output_dir) / "versions"
        versions_dir.mkdir(parents=True, exist_ok=True)
        n = 2
        while (versions_dir / f"Case_01_Phase1_v{n}.xlsx").exists():
            n += 1
        out_path = versions_dir / f"Case_01_Phase1_v{n}.xlsx"

    wb.save(str(out_path))
    logger.info("generate_xlsx: wrote %s", out_path)
    return {"AEGIS-P1-XLSX": str(out_path.resolve())}


# ─────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────


def _build_cover(wb: Workbook, state: Mapping[str, Any]) -> None:
    """Cover sheet — case identification + run metadata."""
    ws = wb.create_sheet("COVER")
    ctx = state.get("company_context")
    ontology = state.get("ontology") or {}
    header = (ontology.get("header") or {}) if isinstance(ontology, Mapping) else {}

    rows: list[tuple[str, str]] = [
        ("Case ID", str(safe_get(header, "case_id", default="-") or "-")),
        ("Phase", str(safe_get(header, "phase", default=1))),
        ("Document Version", str(safe_get(header, "version", default="1.0"))),
        ("Generated Date", str(safe_get(header, "generated_date", default="-"))),
        ("Author", str(safe_get(header, "author", default="-"))),
        ("Company", str(safe_get(ctx, "company_name", default="-") or "-")),
        ("Sector", str(safe_get(ctx, "sector", default="-") or "-")),
        ("Jurisdiction", str(safe_get(ctx, "jurisdiction", default="-") or "-")),
        ("EU Size", str(safe_get(ctx, "scale", default="-") or "-")),
        ("Complexity Tier", str(safe_get(ctx, "complexity_tier", default="-") or "-")),
        ("Pipeline Stage", str(state.get("current_stage", "-"))),
        (
            "# Sub-domains",
            str(
                len(
                    (ontology.get("subdomains") or {}).get("covered", [])
                    if isinstance(ontology.get("subdomains"), Mapping)
                    else []
                )
            ),
        ),
        (
            "# Applicable Regulations",
            str(
                sum(
                    1
                    for r in (state.get("regulations") or ontology.get("regulations") or [])
                    if isinstance(r, Mapping) and r.get("applicable")
                )
            ),
        ),
    ]
    _fill_sheet(ws, ["Field", "Value"], rows)


def _build_company(wb: Workbook, state: Mapping[str, Any]) -> None:
    """COMPANY sheet — derived from ontology.company + company_context."""
    ws = wb.create_sheet("COMPANY")
    ontology = state.get("ontology") or {}
    company = (ontology.get("company") or {}) if isinstance(ontology, Mapping) else {}
    rows: list[tuple[str, str]] = [
        ("ID", str(company.get("id", "-"))),
        ("Name", str(company.get("name", "-"))),
        ("Sector", str(company.get("sector", "-"))),
        ("Size", str(company.get("size", "-"))),
        ("Employees", str(company.get("employees", "-"))),
        (
            "Revenue (EUR)",
            f"{company.get('revenue_eur', 0):,}"
            if isinstance(company.get("revenue_eur"), int | float)
            else "-",
        ),
        ("Jurisdiction", str(company.get("jurisdiction", "-"))),
        ("Legal Structure", str(company.get("legal_structure", "-"))),
        ("Criticality Level", str(company.get("criticality_level", "-"))),
        ("Tech Stack", ", ".join(company.get("tech_stack", []) or [])),
        ("Data Types", ", ".join(company.get("data_types", []) or [])),
    ]
    _fill_sheet(ws, ["Field", "Value"], rows)


def _build_regulations(wb: Workbook, state: Mapping[str, Any]) -> None:
    """REGULATIONS sheet — applicability summary."""
    ws = wb.create_sheet("REGULATIONS")
    regs = state.get("regulations") or (state.get("ontology") or {}).get("regulations", [])
    rows: list[tuple[str, ...]] = []
    for reg in regs:
        if not isinstance(reg, Mapping):
            continue
        rows.append(
            (
                str(reg.get("id", "-")),
                str(reg.get("abbreviation", "-")),
                str(reg.get("name", "-")),
                str(reg.get("eu_reference", "-")),
                "YES" if reg.get("applicable") else "NO",
                _stringify_party(reg.get("obligated_party")),
                str(reg.get("clause_count", 0)),
                str(reg.get("reason", "") or "-"),
            )
        )
    _fill_sheet(
        ws,
        [
            "Reg ID",
            "Abbreviation",
            "Name",
            "EU Reference",
            "Applicable",
            "Obligated Party",
            "Clause Count",
            "Reason",
        ],
        rows,
    )


def _build_clauses(wb: Workbook, state: Mapping[str, Any]) -> None:
    """CLAUSES sheet — clause-to-sub-domain mappings."""
    ws = wb.create_sheet("CLAUSES")
    ontology = state.get("ontology") or {}
    clauses = ontology.get("clause_mappings", []) if isinstance(ontology, Mapping) else []
    regs = ontology.get("regulations", []) if isinstance(ontology, Mapping) else []

    rows: list[tuple[str, ...]] = []
    for clause in clauses:
        if not isinstance(clause, Mapping):
            continue
        rid = str(clause.get("regulation_id", ""))
        abbr = _find_reg_abbreviation(regs, rid)
        rows.append(
            (
                str(clause.get("clause_id", "-")),
                abbr,
                str(clause.get("article", "-")),
                str(clause.get("description", "-")),
                str(clause.get("maps_to_subdomain", "-")),
                str(clause.get("normative_strength", "-")),
                _stringify_party(clause.get("obligated_party", "-")),
                str(clause.get("obligation_type", "-")),
            )
        )
    _fill_sheet(
        ws,
        [
            "Clause ID",
            "Regulation",
            "Article",
            "Description",
            "Sub-domain",
            "Norm. Strength",
            "Obligated Party",
            "Obligation Type",
        ],
        rows,
    )


def _build_coverage(wb: Workbook, state: Mapping[str, Any]) -> None:
    """COVERAGE sheet — sub-domain x regulation coverage matrix."""
    ws = wb.create_sheet("COVERAGE")
    ontology = state.get("ontology") or {}
    subdomains = ontology.get("subdomains", {}) if isinstance(ontology, Mapping) else {}
    regs = state.get("regulations") or ontology.get("regulations", [])

    covered = subdomains.get("covered", []) if isinstance(subdomains, Mapping) else []
    not_covered = subdomains.get("not_covered", []) if isinstance(subdomains, Mapping) else []
    regulations = [r for r in regs if isinstance(r, Mapping)]

    headers = ["Sub-domain", "Name", "State"] + [_abbr(r) for r in regulations]
    rows: list[tuple[str, ...]] = []
    for entry in covered:
        if not isinstance(entry, Mapping):
            continue
        row: list[str] = [
            str(entry.get("id", "-")),
            str(entry.get("name", "-")),
            "COVERED",
        ]
        sources = set(entry.get("source_regulations") or [])
        for reg in regulations:
            row.append("YES" if _abbr(reg) in sources else "NO")
        rows.append(tuple(row))
    for entry in not_covered:
        if not isinstance(entry, Mapping):
            continue
        row = [
            str(entry.get("id", "-")),
            str(entry.get("name", "-")),
            "GAP",
        ]
        # Apply blank Y/N for gaps
        for _ in regulations:
            row.append("NO")
        rows.append(tuple(row))
    _fill_sheet(ws, headers, rows)


def _build_proportionality(wb: Workbook, state: Mapping[str, Any]) -> None:
    """PROPORTIONALITY sheet — TrackB profile (with fallback)."""
    from aegis_phase1.v2.output.doc_07b import _resolve_profile

    profile = _resolve_profile(dict(state))
    ws = wb.create_sheet("PROPORTIONALITY")
    rows: list[tuple[str, ...]] = []
    for sd_id, entry in sorted(profile.items()):
        if not isinstance(entry, Mapping):
            continue
        verification = entry.get("verification_method", ["-"])
        if isinstance(verification, list):
            verification = ", ".join(str(v) for v in verification)
        rows.append(
            (
                str(sd_id),
                str(entry.get("scale", "-")),
                str(entry.get("inheritability", "-")),
                str(entry.get("priority", "-")),
                str(entry.get("tier", "-")),
                str(entry.get("satisfaction_pattern", "-")),
                str(entry.get("evidence_depth", "-")),
                str(verification),
                str(entry.get("ownership", "-")),
                _stringify_party(entry.get("example_controls", ["-"])),
                _stringify_party(entry.get("source_regs", ["-"])),
            )
        )
    _fill_sheet(
        ws,
        [
            "Sub-domain",
            "Scale",
            "Inheritability",
            "Priority",
            "Tier",
            "Satisfaction Pattern",
            "Evidence Depth",
            "Verification Method",
            "Ownership",
            "Example Controls",
            "Source Regs",
        ],
        rows,
    )


def _build_gate(wb: Workbook, state: Mapping[str, Any]) -> None:
    """GATE sheet — completion checklist for Phase 1 → Phase 2."""
    ws = wb.create_sheet("GATE")
    ctx = state.get("company_context")
    ontology = state.get("ontology") or {}
    clauses = ontology.get("clause_mappings", []) if isinstance(ontology, Mapping) else []
    subdomains = ontology.get("subdomains", {}) if isinstance(ontology, Mapping) else {}

    def ok(flag: bool) -> str:
        return "PASS" if flag else "FAIL"

    rows = [
        ("1. Company context loaded", ok(ctx is not None), "see AEGIS-P1-04"),
        (
            "2. Sub-domain catalogue present",
            ok(bool(subdomains)),
            f"{len(subdomains.get('covered', []))} active"
            if isinstance(subdomains, Mapping)
            else "0",
        ),
        (
            "3. Applicable regulations identified",
            ok(len(state.get("regulations") or ontology.get("regulations", [])) > 0),
            f"{len(state.get('regulations') or ontology.get('regulations', []))} regs",
        ),
        ("4. Clause mappings rendered", ok(bool(clauses)), f"{len(clauses)} clauses"),
        ("5. Coverage matrix computed", ok(bool(subdomains)), "see COVERAGE sheet"),
        (
            "6. Proportionality computed",
            ok(bool(state.get("aggregated_data"))),
            "see PROPORTIONALITY sheet",
        ),
    ]
    _fill_sheet(ws, ["#", "Gate criterion", "Status", "Evidence"], rows)


# ─────────────────────────────────────────────────────────────────────
# Style / layout helpers
# ─────────────────────────────────────────────────────────────────────


def _fill_sheet(ws: Worksheet, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    """Populate a worksheet with ``headers`` + ``rows`` and apply formatting."""
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    for idx, row in enumerate(rows, start=2):
        ws.append([_cell(value) for value in row])
        if idx % 2 == 0:
            for cell in ws[idx]:
                cell.fill = _ZEBRA_FILL

    widths = _column_widths(headers, rows)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 80)


def _cell(value: Any) -> Any:
    """Convert cells so openpyxl does not write ``None`` literals."""
    if value is None:
        return "-"
    if isinstance(value, str | int | float | bool):
        return value
    return str(value)


def _column_widths(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> list[int]:
    """Return the maximum width per column (in characters)."""
    widths: list[int] = [len(str(h)) + 2 for h in headers]
    for row in rows:
        for i, value in enumerate(row):
            text = str(value) if value is not None else ""
            widths[i] = max(widths[i], len(text) + 2)
    return widths


def _stringify_party(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if value is None:
        return "-"
    return str(value)


def _abbr(reg: Mapping[str, Any]) -> str:
    return str(reg.get("abbreviation") or reg.get("id") or "?")


def _find_reg_abbreviation(regs: list[Any], reg_id: str) -> str:
    for reg in regs:
        if isinstance(reg, Mapping) and reg.get("id") == reg_id:
            abbr = reg.get("abbreviation")
            if abbr:
                return str(abbr)
            return str(reg_id).split("/")[-1]
    return str(reg_id).split("/")[-1]


__all__ = ["generate_xlsx"]
